from flask import Flask, render_template, request, redirect, session, jsonify, send_file, url_for
from model import predict_theft
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
import sqlite3
from functools import wraps
from datetime import datetime
import os

# Load .env file if present (for local development)
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv not installed; use system environment variables

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "electroguard_maha_2024_secret_v2")

def get_db():
    conn = sqlite3.connect("electroguard.db")
    conn.row_factory = sqlite3.Row
    return conn

def init_database():
    conn = get_db()
    c = conn.cursor()
    c.execute("""CREATE TABLE IF NOT EXISTS users (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        username TEXT UNIQUE NOT NULL,
        password TEXT NOT NULL,
        role TEXT NOT NULL CHECK(role IN ('state','district','village')),
        district TEXT, village TEXT, full_name TEXT, email TEXT, phone TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
        last_login TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS households (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        house_no TEXT NOT NULL, consumer_name TEXT, people INTEGER, appliances INTEGER,
        expected_usage REAL, district TEXT, village TEXT, meter_id TEXT,
        meter_status TEXT DEFAULT 'active',
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS theft_reports (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        house_no TEXT, consumer_name TEXT,
        voltage REAL, current REAL, power REAL, time REAL, meter_diff REAL,
        result TEXT, confidence REAL, severity TEXT, theft_type TEXT, reasons TEXT,
        reported_by TEXT, district TEXT, village TEXT,
        meter_status TEXT DEFAULT 'active',
        fir_status TEXT DEFAULT 'none',
        fir_number TEXT, fir_filed_by TEXT, fir_filed_at TIMESTAMP, fir_notes TEXT,
        meter_cut_by TEXT, meter_cut_at TIMESTAMP, meter_cut_notes TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS activity_log (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        action TEXT NOT NULL, description TEXT, performed_by TEXT,
        role TEXT, district TEXT, village TEXT, report_id INTEGER,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""CREATE TABLE IF NOT EXISTS notifications (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        title TEXT, message TEXT, type TEXT DEFAULT 'info',
        target_role TEXT DEFAULT 'all', target_district TEXT,
        is_read INTEGER DEFAULT 0, created_by TEXT,
        created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP)""")
    c.execute("""INSERT OR IGNORE INTO users (username,password,role,full_name,email)
        VALUES (?,?,?,?,?)""",
        ("superadmin","admin123","state","Maharashtra State Admin","admin@mahaelectro.gov.in"))
    conn.commit()
    conn.close()

init_database()

BASE = os.path.dirname(os.path.abspath(__file__))
district_df = pd.read_csv(os.path.join(BASE,"districtofSpecificState20191224033051634.csv"), on_bad_lines='skip', encoding='latin1')
village_df  = pd.read_csv(os.path.join(BASE,"villageofSpecificState20191224033059092.csv"),  on_bad_lines='skip', encoding='latin1')
district_col  = [c for c in district_df.columns if "District Name" in c][0]
vill_dist_col = [c for c in village_df.columns if "District Name" in c][0]
vill_name_col = [c for c in village_df.columns if "Village Name" in c][0]

def log_activity(action, description, report_id=None):
    try:
        conn = get_db()
        conn.execute("""INSERT INTO activity_log
            (action,description,performed_by,role,district,village,report_id) VALUES (?,?,?,?,?,?,?)""",
            (action, description, session.get("user","system"), session.get("role",""),
             session.get("district",""), session.get("village",""), report_id))
        conn.commit()
        conn.close()
    except: pass

def get_unread_notifications():
    try:
        conn = get_db()
        role = session.get("role","")
        district = session.get("district","")
        notifs = conn.execute("""SELECT * FROM notifications WHERE is_read=0
            AND (target_role='all' OR target_role=? OR (target_role='district' AND target_district=?))
            ORDER BY created_at DESC LIMIT 10""", (role, district)).fetchall()
        conn.close()
        return list(notifs), len(notifs)
    except: return [], 0

def login_required(f):
    @wraps(f)
    def wrapped(*args, **kwargs):
        if "user" not in session: return redirect(url_for("login"))
        return f(*args, **kwargs)
    return wrapped

def role_required(roles):
    def decorator(f):
        @wraps(f)
        def wrapped(*args, **kwargs):
            if "user" not in session: return redirect(url_for("login"))
            if session.get("role") not in roles:
                return render_template("error.html", msg="Access Denied"), 403
            return f(*args, **kwargs)
        return wrapped
    return decorator

@app.route("/")
def home():
    return redirect(url_for("dashboard") if "user" in session else url_for("login"))

@app.route("/login", methods=["GET","POST"])
def login():
    if "user" in session: return redirect(url_for("dashboard"))
    error = None
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"]
        conn = get_db()
        user = conn.execute("SELECT * FROM users WHERE username=? AND password=?", (username,password)).fetchone()
        if user:
            conn.execute("UPDATE users SET last_login=? WHERE id=?", (datetime.now(), user["id"]))
            conn.commit()
            conn.close()
            session.update({"user":user["username"],"user_id":user["id"],"role":user["role"],
                "district":user["district"],"village":user["village"],"full_name":user["full_name"] or user["username"]})
            log_activity("LOGIN", f"Logged in from {request.remote_addr}")
            return redirect(url_for("dashboard"))
        conn.close()
        error = "Invalid username or password"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    log_activity("LOGOUT","User logged out")
    session.clear()
    return redirect(url_for("login"))

@app.route("/dashboard")
@login_required
def dashboard():
    conn = get_db()
    role = session["role"]
    stats = {
        "total_admins":    conn.execute("SELECT COUNT(*) FROM users").fetchone()[0],
        "district_admins": conn.execute("SELECT COUNT(*) FROM users WHERE role='district'").fetchone()[0],
        "village_admins":  conn.execute("SELECT COUNT(*) FROM users WHERE role='village'").fetchone()[0],
    }
    if role == "state":
        houses = conn.execute("SELECT * FROM households ORDER BY id DESC LIMIT 50").fetchall()
        stats.update({
            "total_households": conn.execute("SELECT COUNT(*) FROM households").fetchone()[0],
            "theft_cases":  conn.execute("SELECT COUNT(*) FROM theft_reports WHERE result='THEFT_DETECTED'").fetchone()[0],
            "safe_cases":   conn.execute("SELECT COUNT(*) FROM theft_reports WHERE result='SAFE'").fetchone()[0],
            "fir_filed":    conn.execute("SELECT COUNT(*) FROM theft_reports WHERE fir_status='filed'").fetchone()[0],
            "meters_cut":   conn.execute("SELECT COUNT(*) FROM theft_reports WHERE meter_status='cut'").fetchone()[0],
        })
        recent_reports  = conn.execute("SELECT * FROM theft_reports ORDER BY created_at DESC LIMIT 8").fetchall()
        critical_cases  = conn.execute("SELECT * FROM theft_reports WHERE result='THEFT_DETECTED' AND meter_status='active' ORDER BY created_at DESC LIMIT 5").fetchall()
    elif role == "district":
        d = session["district"]
        houses = conn.execute("SELECT * FROM households WHERE district=? ORDER BY id DESC LIMIT 50",(d,)).fetchall()
        stats.update({
            "total_households": conn.execute("SELECT COUNT(*) FROM households WHERE district=?",(d,)).fetchone()[0],
            "theft_cases":  conn.execute("SELECT COUNT(*) FROM theft_reports WHERE result='THEFT_DETECTED' AND district=?",(d,)).fetchone()[0],
            "safe_cases":   conn.execute("SELECT COUNT(*) FROM theft_reports WHERE result='SAFE' AND district=?",(d,)).fetchone()[0],
            "fir_filed":    conn.execute("SELECT COUNT(*) FROM theft_reports WHERE fir_status='filed' AND district=?",(d,)).fetchone()[0],
            "meters_cut":   conn.execute("SELECT COUNT(*) FROM theft_reports WHERE meter_status='cut' AND district=?",(d,)).fetchone()[0],
        })
        recent_reports = conn.execute("SELECT * FROM theft_reports WHERE district=? ORDER BY created_at DESC LIMIT 8",(d,)).fetchall()
        critical_cases = conn.execute("SELECT * FROM theft_reports WHERE result='THEFT_DETECTED' AND meter_status='active' AND district=? ORDER BY created_at DESC LIMIT 5",(d,)).fetchall()
    else:
        v,d = session["village"], session["district"]
        houses = conn.execute("SELECT * FROM households WHERE village=? AND district=? ORDER BY id DESC LIMIT 50",(v,d)).fetchall()
        stats.update({
            "total_households": conn.execute("SELECT COUNT(*) FROM households WHERE village=? AND district=?",(v,d)).fetchone()[0],
            "theft_cases":  conn.execute("SELECT COUNT(*) FROM theft_reports WHERE result='THEFT_DETECTED' AND village=?",(v,)).fetchone()[0],
            "safe_cases":   conn.execute("SELECT COUNT(*) FROM theft_reports WHERE result='SAFE' AND village=?",(v,)).fetchone()[0],
            "fir_filed":    conn.execute("SELECT COUNT(*) FROM theft_reports WHERE fir_status='filed' AND village=?",(v,)).fetchone()[0],
            "meters_cut":   conn.execute("SELECT COUNT(*) FROM theft_reports WHERE meter_status='cut' AND village=?",(v,)).fetchone()[0],
        })
        recent_reports = conn.execute("SELECT * FROM theft_reports WHERE village=? ORDER BY created_at DESC LIMIT 8",(v,)).fetchall()
        critical_cases = conn.execute("SELECT * FROM theft_reports WHERE result='THEFT_DETECTED' AND meter_status='active' AND village=? ORDER BY created_at DESC LIMIT 5",(v,)).fetchall()
    notifs, notif_count = get_unread_notifications()
    conn.close()
    return render_template("dashboard.html", stats=stats, houses=houses,
        recent_reports=recent_reports, critical_cases=critical_cases,
        notifs=notifs, notif_count=notif_count)

@app.route("/predict_page")
@login_required
def predict_page():
    notifs, notif_count = get_unread_notifications()
    return render_template("predict.html", notifs=notifs, notif_count=notif_count)

@app.route("/predict", methods=["POST"])
@login_required
def predict():
    try:
        voltage    = float(request.form["voltage"])
        current    = float(request.form["current"])
        power      = float(request.form["power"])
        time       = float(request.form["time"])
        meter_diff = float(request.form["meter_diff"])
        house_no   = request.form["house_no"].strip()
        consumer   = request.form["consumer_name"].strip()
        result, confidence, reasons, theft_type, severity = predict_theft(voltage, current, power, time, meter_diff)
        conn = get_db()
        conn.execute("""INSERT INTO theft_reports
            (house_no,consumer_name,voltage,current,power,time,meter_diff,
             result,confidence,severity,theft_type,reasons,reported_by,district,village)
            VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
            (house_no, consumer, voltage, current, power, time, meter_diff,
             result, confidence, severity, theft_type, "; ".join(reasons),
             session["user"], session.get("district",""), session.get("village","")))
        report_id = conn.execute("SELECT last_insert_rowid()").fetchone()[0]
        if result == "THEFT_DETECTED" and severity == "CRITICAL":
            conn.execute("""INSERT INTO notifications
                (title,message,type,target_role,target_district,created_by) VALUES (?,?,?,?,?,?)""",
                (f"CRITICAL Theft — {house_no}",
                 f"Consumer: {consumer} | By: {session['user']} | District: {session.get('district','')}",
                 "danger","state",session.get("district",""),session["user"]))
        conn.commit()
        conn.close()
        log_activity("PREDICTION", f"Detection for {house_no} ({consumer}) — {result}", report_id)
        session["last_report"] = {"id":report_id,"house_no":house_no,"consumer":consumer,
            "voltage":voltage,"current":current,"power":power,"time":time,"meter_diff":meter_diff,
            "result":result,"confidence":confidence,"reasons":reasons,"theft_type":theft_type,"severity":severity}
        notifs, notif_count = get_unread_notifications()
        return render_template("predict.html", result=result, confidence=confidence, reasons=reasons,
            theft_type=theft_type, severity=severity, house_no=house_no, consumer=consumer,
            voltage=voltage, current=current, power=power, time=time, meter_diff=meter_diff,
            report_id=report_id, notifs=notifs, notif_count=notif_count)
    except Exception as e:
        notifs, notif_count = get_unread_notifications()
        return render_template("predict.html", error=str(e), notifs=notifs, notif_count=notif_count)

@app.route("/cut_meter/<int:report_id>", methods=["POST"])
@login_required
def cut_meter(report_id):
    notes = request.form.get("notes","").strip()
    conn = get_db()
    report = conn.execute("SELECT * FROM theft_reports WHERE id=?", (report_id,)).fetchone()
    if not report:
        conn.close()
        return jsonify({"success":False,"msg":"Report not found"}), 404
    if report["meter_status"] == "cut":
        conn.close()
        return jsonify({"success":False,"msg":"Meter already disconnected"})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("""UPDATE theft_reports SET meter_status='cut',meter_cut_by=?,meter_cut_at=?,meter_cut_notes=?
        WHERE id=?""", (session["user"], now, notes, report_id))
    conn.execute("""INSERT INTO notifications (title,message,type,target_role,target_district,created_by)
        VALUES (?,?,?,?,?,?)""",
        (f"Meter Cut — {report['house_no']}",
         f"Consumer: {report['consumer_name']} | By: {session['user']} | {report['district']} > {report['village']}",
         "warning","all",report["district"],session["user"]))
    conn.commit()
    conn.close()
    log_activity("CUT_METER", f"Meter disconnected: {report['house_no']} ({report['consumer_name']}). Notes: {notes}", report_id)
    return jsonify({"success":True,
        "msg":f"Meter for {report['house_no']} disconnected successfully.",
        "cut_by":session["user"],
        "cut_at":datetime.now().strftime("%d %b %Y, %H:%M")})

@app.route("/file_fir/<int:report_id>", methods=["POST"])
@login_required
def file_fir(report_id):
    fir_number = request.form.get("fir_number","").strip()
    notes      = request.form.get("notes","").strip()
    if not fir_number:
        return jsonify({"success":False,"msg":"FIR number is required"})
    conn = get_db()
    report = conn.execute("SELECT * FROM theft_reports WHERE id=?", (report_id,)).fetchone()
    if not report:
        conn.close()
        return jsonify({"success":False,"msg":"Report not found"}), 404
    if report["fir_status"] == "filed":
        conn.close()
        return jsonify({"success":False,"msg":"FIR already filed for this case"})
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    conn.execute("""UPDATE theft_reports SET fir_status='filed',fir_number=?,fir_filed_by=?,fir_filed_at=?,fir_notes=?
        WHERE id=?""", (fir_number, session["user"], now, notes, report_id))
    conn.execute("""INSERT INTO notifications (title,message,type,target_role,target_district,created_by)
        VALUES (?,?,?,?,?,?)""",
        (f"FIR Filed — {report['house_no']} | #{fir_number}",
         f"Consumer: {report['consumer_name']} | FIR #{fir_number} | Filed by: {session['user']}",
         "info","state",report["district"],session["user"]))
    conn.commit()
    conn.close()
    log_activity("FILE_FIR", f"FIR #{fir_number} for {report['house_no']} ({report['consumer_name']}). Notes: {notes}", report_id)
    return jsonify({"success":True,
        "msg":f"FIR #{fir_number} filed successfully.",
        "fir_number":fir_number,
        "filed_by":session["user"],
        "filed_at":datetime.now().strftime("%d %b %Y, %H:%M")})

@app.route("/admin", methods=["GET","POST"])
@role_required(["state","district"])
def admin_panel():
    conn = get_db()
    role = session["role"]
    msg = None
    if request.method == "POST":
        action = request.form.get("action")
        if action == "create_district" and role == "state":
            uname = request.form["username"].strip()
            try:
                conn.execute("""INSERT INTO users (username,password,role,district,full_name,email)
                    VALUES (?,?,?,?,?,?)""",
                    (uname, request.form["password"], "district", request.form["district_name"],
                     request.form.get("full_name",""), request.form.get("email","")))
                conn.commit()
                log_activity("CREATE_ADMIN", f"Created district admin '{uname}'")
                msg = ("success", f"District admin '{uname}' created!")
            except: msg = ("error", f"Username '{uname}' already exists.")
        elif action == "create_village":
            uname = request.form["username"].strip()
            dist  = request.form.get("district_name", session.get("district",""))
            if role == "district": dist = session["district"]
            try:
                conn.execute("""INSERT INTO users (username,password,role,district,village,full_name)
                    VALUES (?,?,?,?,?,?)""",
                    (uname, request.form["password"], "village", dist,
                     request.form["village_name"], request.form.get("full_name","")))
                conn.commit()
                log_activity("CREATE_ADMIN", f"Created village admin '{uname}'")
                msg = ("success", f"Village admin '{uname}' created!")
            except: msg = ("error", f"Username '{uname}' already exists.")
        elif action == "add_household":
            try:
                conn.execute("""INSERT INTO households
                    (house_no,consumer_name,people,appliances,expected_usage,district,village,meter_id)
                    VALUES (?,?,?,?,?,?,?,?)""",
                    (request.form["house_no"], request.form.get("consumer_name",""),
                     int(request.form.get("people",1)), int(request.form.get("appliances",1)),
                     float(request.form.get("expected_usage",0)),
                     request.form["district"], request.form["village"], request.form.get("meter_id","")))
                conn.commit()
                log_activity("ADD_HOUSEHOLD", f"Added household {request.form['house_no']}")
                msg = ("success", "Household added successfully!")
            except Exception as e: msg = ("error", str(e))
    district_admins = conn.execute("SELECT * FROM users WHERE role='district' ORDER BY district").fetchall()
    if role == "state":
        village_admins = conn.execute("SELECT * FROM users WHERE role='village' ORDER BY district,village").fetchall()
    else:
        village_admins = conn.execute("SELECT * FROM users WHERE role='village' AND district=? ORDER BY village",(session["district"],)).fetchall()
    notifs, notif_count = get_unread_notifications()
    conn.close()
    return render_template("admin.html", role=role, district_admins=district_admins,
        village_admins=village_admins, msg=msg, notifs=notifs, notif_count=notif_count)

@app.route("/delete_admin/<int:uid>")
@role_required(["state","district"])
def delete_admin(uid):
    conn = get_db()
    user = conn.execute("SELECT * FROM users WHERE id=?",(uid,)).fetchone()
    if user and user["role"] != "state":
        if session["role"] == "state" or (session["role"] == "district" and user["district"] == session["district"]):
            conn.execute("DELETE FROM users WHERE id=?",(uid,))
            conn.commit()
            log_activity("DELETE_ADMIN", f"Deleted admin '{user['username']}'")
    conn.close()
    return redirect(url_for("admin_panel"))

@app.route("/reports")
@login_required
def reports():
    conn = get_db()
    role = session["role"]
    severity_filter = request.args.get("severity","all")
    result_filter   = request.args.get("result","all")
    base = "SELECT * FROM theft_reports WHERE 1=1"
    params = []
    if role == "district": base += " AND district=?"; params.append(session["district"])
    elif role == "village": base += " AND village=?"; params.append(session["village"])
    if severity_filter != "all": base += " AND severity=?"; params.append(severity_filter.upper())
    if result_filter   != "all": base += " AND result=?";   params.append(result_filter.upper())
    base += " ORDER BY created_at DESC"
    all_reports = conn.execute(base, params).fetchall()
    notifs, notif_count = get_unread_notifications()
    conn.close()
    return render_template("reports.html", reports=all_reports,
        severity_filter=severity_filter, result_filter=result_filter,
        notifs=notifs, notif_count=notif_count)

@app.route("/activity")
@role_required(["state","district"])
def activity_log_view():
    conn = get_db()
    if session["role"] == "state":
        logs = conn.execute("SELECT * FROM activity_log ORDER BY created_at DESC LIMIT 200").fetchall()
    else:
        logs = conn.execute("SELECT * FROM activity_log WHERE district=? ORDER BY created_at DESC LIMIT 100",(session["district"],)).fetchall()
    notifs, notif_count = get_unread_notifications()
    conn.close()
    return render_template("activity.html", logs=logs, notifs=notifs, notif_count=notif_count)

@app.route("/mark_notifications_read", methods=["POST"])
@login_required
def mark_notifs_read():
    conn = get_db()
    conn.execute("UPDATE notifications SET is_read=1")
    conn.commit()
    conn.close()
    return jsonify({"success":True})

@app.route("/download_report")
@login_required
def download_report():
    rpt = session.get("last_report")
    if not rpt:
        return redirect(url_for("predict_page"))

    conn = get_db()
    db_rpt = conn.execute(
        "SELECT * FROM theft_reports WHERE id=?",
        (rpt.get("id", 0),)
    ).fetchone() if rpt.get("id") else None
    conn.close()

    wb = Workbook()
    ws = wb.active
    ws.title = "ElectroGuard Report"

    # ================= HEADER =================
    ws.merge_cells("A1:L1")
    ws["A1"] = "ELECTROGUARD — Electricity Theft Detection Report"
    ws["A1"].font = Font(bold=True, size=13, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="03102B")
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:L2")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')} | By: {session['user']} | Role: {session['role'].upper()}"
    ws["A2"].alignment = Alignment(horizontal="center")

    ws.row_dimensions[1].height = 28
    ws.row_dimensions[2].height = 20

    ws.append([])

    # ================= TABLE =================
    headers = [
        "House No","Consumer","Voltage(V)","Current(A)","Power(W)",
        "Time(h)","Meter Diff(kWh)","Result","Confidence%",
        "Severity","Theft Type","Reasons"
    ]

    hfill = PatternFill("solid", fgColor="0B3D91")
    hfont = Font(color="FFFFFF", bold=True)

    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=i, value=h)
        cell.fill = hfill
        cell.font = hfont
        cell.alignment = Alignment(horizontal="center")

    ws.append([
        rpt["house_no"],
        rpt["consumer"],
        rpt["voltage"],
        rpt["current"],
        rpt["power"],
        rpt["time"],
        rpt["meter_diff"],
        "THEFT DETECTED" if rpt["result"] == "THEFT_DETECTED" else "SAFE",
        rpt["confidence"],
        rpt.get("severity", ""),
        rpt.get("theft_type", "N/A"),
        "; ".join(rpt.get("reasons", []))
    ])

    # ================= SMART COLOR =================
    if rpt["result"] == "THEFT_DETECTED":
        color = "FFCCCC"  # red
    else:
        color = "CCFFCC"  # green

    for col in range(1, len(headers) + 1):
        ws.cell(row=5, column=col).fill = PatternFill("solid", fgColor=color)

    # ================= BORDERS =================
    from openpyxl.styles import Border, Side

    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for row in ws.iter_rows(min_row=4, max_row=5, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.row_dimensions[4].height = 22
    ws.row_dimensions[5].height = 20

    # ================= SEVERITY HIGHLIGHT =================
    severity_col = headers.index("Severity") + 1
    severity_cell = ws.cell(row=5, column=severity_col)

    if rpt.get("severity") == "CRITICAL":
        severity_cell.font = Font(color="FF0000", bold=True)
    elif rpt.get("severity") == "MODERATE":
        severity_cell.font = Font(color="FF8C00", bold=True)

    # ================= ACTION STATUS =================
    ws.append([])
    ws.append(["=== ACTION STATUS ==="])
    ws["A7"].font = Font(bold=True, size=11, color="0B3D91")

    if db_rpt:
        ws.append([
            "Meter Status:", db_rpt["meter_status"].upper(),
            "Cut By:", db_rpt["meter_cut_by"] or "—",
            "Cut At:", db_rpt["meter_cut_at"] or "—"
        ])

        ws.append([
            "FIR Status:", db_rpt["fir_status"].upper(),
            "FIR No.:", db_rpt["fir_number"] or "—",
            "Filed By:", db_rpt["fir_filed_by"] or "—",
            "Filed At:", db_rpt["fir_filed_at"] or "—"
        ])

        if db_rpt["fir_notes"]:
            ws.append(["FIR Notes:", db_rpt["fir_notes"]])

        if db_rpt["meter_cut_notes"]:
            ws.append(["Cut Notes:", db_rpt["meter_cut_notes"]])

    # ================= COLUMN WIDTH =================
    from openpyxl.utils import get_column_letter

    for i in range(1, ws.max_column + 1):
        col_letter = get_column_letter(i)
        ws.column_dimensions[col_letter].width = 22

    # ================= SAVE =================
    filename = f"ElectroGuard_{rpt['house_no']}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    os.makedirs("exports", exist_ok=True)
    path = os.path.join("exports", filename)

    wb.save(path)

    return send_file(
        path,
        as_attachment=True,
        download_name=filename
    )
@app.route("/settings", methods=["GET","POST"])
@login_required
def settings():
    conn = get_db()
    msg = None
    if request.method == "POST":
        action = request.form.get("action")
        if action == "change_password":
            old = request.form["old_password"]; new = request.form["new_password"]
            user = conn.execute("SELECT * FROM users WHERE id=? AND password=?",(session["user_id"],old)).fetchone()
            if user:
                conn.execute("UPDATE users SET password=? WHERE id=?",(new,session["user_id"]))
                conn.commit(); log_activity("CHANGE_PASSWORD","Password changed")
                msg = ("success","Password changed successfully!")
            else: msg = ("error","Current password is incorrect.")
        elif action == "update_profile":
            fname=request.form.get("full_name",""); email=request.form.get("email",""); phone=request.form.get("phone","")
            conn.execute("UPDATE users SET full_name=?,email=?,phone=? WHERE id=?",(fname,email,phone,session["user_id"]))
            conn.commit(); session["full_name"]=fname; log_activity("UPDATE_PROFILE","Profile updated")
            msg = ("success","Profile updated!")
    user_data = conn.execute("SELECT * FROM users WHERE id=?",(session["user_id"],)).fetchone()
    notifs, notif_count = get_unread_notifications()
    conn.close()
    return render_template("settings.html", user_data=user_data, msg=msg, notifs=notifs, notif_count=notif_count)

@app.route("/get_districts")
def get_districts():
    districts = sorted(district_df[district_col].dropna().unique().tolist())
    return jsonify({"districts": districts})

@app.route("/get_villages/<district>")
def get_villages(district):
    district = district.strip().upper()
    vdf = village_df.copy()
    vdf[vill_dist_col] = vdf[vill_dist_col].astype(str).str.strip().str.upper()
    vdf[vill_name_col] = vdf[vill_name_col].astype(str).str.strip()
    filtered = vdf[vdf[vill_dist_col] == district]
    villages = sorted(filtered[vill_name_col].dropna().unique().tolist())
    return jsonify({"villages": villages[:300]})

@app.route("/api/report/<int:report_id>")
@login_required
def api_report(report_id):
    conn = get_db()
    r = conn.execute("SELECT * FROM theft_reports WHERE id=?",(report_id,)).fetchone()
    conn.close()
    if not r: return jsonify({"error":"not found"}), 404
    return jsonify({"id":r["id"],"house_no":r["house_no"],"consumer_name":r["consumer_name"],
        "meter_status":r["meter_status"],"fir_status":r["fir_status"],"fir_number":r["fir_number"],
        "meter_cut_by":r["meter_cut_by"],"meter_cut_at":r["meter_cut_at"],
        "fir_filed_by":r["fir_filed_by"],"fir_filed_at":r["fir_filed_at"]})

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
